"""Утилиты для экспорта данных в Excel/CSV."""
import io
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from services.db import ProductRow
import csv


async def generate_excel(products: List[ProductRow], discount: int = 0) -> io.BytesIO:
    """
    Генерирует Excel файл со списком товаров.
    
    Args:
        products: Список товаров
        discount: Процент скидки WB кошелька
        
    Returns:
        BytesIO объект с Excel файлом
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары WB"
    
    # Заголовки
    headers = ["№", "Название", "Артикул", "Размер", "Цена", "С кошельком", "Остаток", "Ссылка"]
    ws.append(headers)
    
    # Стилизация заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Данные
    for idx, product in enumerate(products, 1):
        price = product.last_product_price or 0
        price_with_wallet = price
        
        if discount > 0:
            from utils.wb_utils import apply_wallet_discount
            price_with_wallet = apply_wallet_discount(price, discount)
        
        size = product.selected_size if product.selected_size else "—"
        stock = product.last_qty if product.last_qty is not None else "—"
        
        if product.out_of_stock:
            stock = "Нет в наличии"
        
        row_data = [
            idx,
            product.display_name,
            product.nm_id,
            size,
            f"{price} ₽",
            f"{price_with_wallet} ₽" if discount > 0 else "—",
            stock,
            product.url_product
        ]
        
        ws.append(row_data)
    
    # Автоширина столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в BytesIO
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


async def generate_csv(products: List[ProductRow], discount: int = 0) -> io.BytesIO:
    """
    Генерирует CSV файл со списком товаров.
    
    Args:
        products: Список товаров
        discount: Процент скидки WB кошелька
        
    Returns:
        BytesIO объект с CSV файлом
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=';')
    
    # Заголовки
    headers = ["№", "Название", "Артикул", "Размер", "Цена", "С кошельком", "Остаток", "Ссылка"]
    writer.writerow(headers)
    
    # Данные
    for idx, product in enumerate(products, 1):
        price = product.last_product_price or 0
        price_with_wallet = price
        
        if discount > 0:
            from utils.wb_utils import apply_wallet_discount
            price_with_wallet = apply_wallet_discount(price, discount)
        
        size = product.selected_size if product.selected_size else "—"
        stock = product.last_qty if product.last_qty is not None else "—"
        
        if product.out_of_stock:
            stock = "Нет в наличии"
        
        row_data = [
            idx,
            product.display_name,
            product.nm_id,
            size,
            f"{price} ₽",
            f"{price_with_wallet} ₽" if discount > 0 else "—",
            stock,
            product.url_product
        ]
        
        writer.writerow(row_data)
    
    # Конвертируем в BytesIO
    output = io.BytesIO()
    output.write(buffer.getvalue().encode('utf-8-sig'))  # BOM для корректного отображения в Excel
    output.seek(0)
    
    return output
